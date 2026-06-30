# excel_engine.py
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel_report(dataframe):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ecosystem Data"
    ws.sheet_view.showGridLines = True
    
    title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
    )
    
    ws.merge_cells("A1:G1")
    ws["A1"] = "SANDHAR ECOSYSTEM DATA MANAGEMENT REPORT"
    ws["A1"].font = font_title
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions.height = 35
    
    cols = ["Vertical Segment", "Plant Node", "Location", "Grid (MVAh)", "Mitigation (MT)", "Emissions (MT)", "Yield Ratio"]
    for idx, text in enumerate(cols, 1):
        cell = ws.cell(row=3, column=idx, value=text)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions.height = 24
    
    for r_idx, row in dataframe.reset_index().iterrows():
        r = r_idx + 4
        ws.cell(row=r, column=1, value=row['vertical']).font = font_data
        ws.cell(row=r, column=2, value=row['unit']).font = font_data
        ws.cell(row=r, column=3, value=row['location']).font = font_data
        
        c4 = ws.cell(row=r, column=4, value=row['grid_mvah'])
        c4.number_format = "#,##0.00"
        c4.font = font_data
        
        c5 = ws.cell(row=r, column=5, value=row['mitigation'])
        c5.number_format = "#,##0"
        c5.font = font_data
        
        c6 = ws.cell(row=r, column=6, value=row['emission'])
        c6.number_format = "#,##0"
        c6.font = font_data
        
        c7 = ws.cell(row=r, column=7, value=row['generation_per_kwp'])
        c7.number_format = "0.00"
        c7.font = font_data
        
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col.column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output)
    return output.getvalue()
