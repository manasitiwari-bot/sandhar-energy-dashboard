import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -----------------------------
# Sample Data (Generation per KWP row)
# -----------------------------
data = {
    "Unit": ["SAG", "SCD", "SEB", "SAD", "SCR", "STPL"],
    "Generation per KWP": [3.28, 2.75, 4.28, 1.16, 4.92, 2.64]
}

df = pd.DataFrame(data)

# -----------------------------
# Save to Excel
# -----------------------------
file_name = "generation_kwp.xlsx"
df.to_excel(file_name, index=False)

# -----------------------------
# Apply Conditional Formatting
# -----------------------------
wb = load_workbook(file_name)
ws = wb.active

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for row in range(2, ws.max_row + 1):
    cell = ws[f"B{row}"]
    if cell.value < 3:
        cell.fill = red_fill
    else:
        cell.fill = green_fill

wb.save(file_name)

print("✅ Excel file saved with conditional formatting")
