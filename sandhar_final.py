import streamlit as st
import pandas as pd
import io
import plotly.express as px
import plotly.graph_objects as go
import folium
from streamlit_folium import st_folium
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 1. Page Configuration
st.set_page_config(
    page_title="Sandhar Energy Ecosystem Dashboard",
    page_icon="🌱",
    layout="wide"
)

# --- GLOBAL STATIC MULTILINE CSV DATA MATRICES ---
MASTER_CSV_DATA = """vertical,unit,location,grid_mvah,capex_capacity,opex_capacity,replacement_pct,dg,mitigation,emission,capex_gen,opex_gen,lat,lon,unit_lost_inefficiency,generation_per_kwp,gen_kwp_27
Automotive Business,SAG,Gurugram,2074.867,33,0,34,17386,515,1508,33.878,0.0,28.4595,77.0266,4933,2.81,3.16
Plastic Business,SCD,Gurugram,2538.911,110,132,42,52045,772,1846,0.0,237.971,28.4595,77.0266,54539,1.93,2.92
Sheet Metal & Allied Business,SEB,Gurugram,2955.38,50,300,12,0,265,2149,364.741,0.0,28.4595,77.0266,0,3.28,4.50
Automotive Business,SAD,Gurugram,4614.016,138,218,34,43194,1139,3354,36.582,106.451,28.4595,77.0266,54909,2.86,3.88
Casting Machining & Tooling Business,SCR,Gurugram,12081.709,50,0,30,5730,2653,8783,0.0,8.865,28.4595,77.0266,24938,2.79,4.02
Casting Machining & Tooling Business,STPL,Gurugram,394.008,36,0,8,55432,24,286,0.0,33.266,28.4595,77.0266,0,2.20,4.56
Casting Machining & Tooling Business,ACM,Gurugram,3249.812,127,0,22,2200,525,2363,0.0,122.96,28.4595,77.0266,51085,0.49,1.26
Corp. Office,CORP,Gurugram,232.695,25,0,14,0,24,169,0.0,33.483,28.4595,77.0266,9898,2.53,3.46
Corp. Office,SASPL,Tamil Nadu,159.629,0,0,41,0,48,116,0.0,66.246,11.1271,78.6569,29313,2.65,3.71
Automotive Business,SHP,Rajasthan,881.47,400,262,60,13021,381,641,346.631,1779.14,27.0238,74.2179,113748,3.67,5.02
Automotive Business,SAH,Uttarakhand,5657.6,251,129,5,66015,17,4113,104.663,166.675,30.0668,79.0193,167105,1.21,0.00
Casting Machining & Tooling Business,SCH,Tamil Nadu,18864.562,0,0,53,0,7304,13715,3979.8,0.0,11.1271,78.6569,0,2.14,2.71
Automotive Business,SIO,Tamil Nadu,1271.13,125,0,9,15220,81,924,110.985,0.0,11.1271,78.6569,0,2.70,5.20
Casting Machining & Tooling Business,SCA,Karnataka,4232.325,115,0,59,1111,1825,3077,0.0,88.659,15.3173,75.7139,45077,2.79,2.38
Automotive Business,SAB,Karnataka,1779.327,0,340,93,10120,1204,1294,442.802,0.0,15.3173,75.7139,0,2.31,4.90
Sheet Metal & Allied Business,SCY,Karnataka,2695.544,0,634,33,0,642,1960,883.541,0.0,15.3173,75.7139,0,3.05,3.46
Cabin & Fabrication Division,SIP,Pune,377.270,717,0,236,3490,648,274,891.451,0.0,18.5204,73.8567,13106,2.77,2.62
Cabin & Fabrication Division,SIA,Karnataka,1506.093,115,0,2,6470,22,1095,0.0,29.61,15.3173,75.7139,39572,2.11,1.67
Casting Machining & Tooling Business,SKC,Pune,1653.615,0,604,23,11891,277,1202,381.131,381.131,18.5204,73.8567,0,3.55,4.34
Sheet Metal & Allied Business,SHN,Tamil Nadu,2079.137,0,624,19,15522,290,1512,398.942,0.0,11.1271,78.6569,0,3.66,4.34"""

MONTHLY_CSV_DATA = """Month,SAG,SCD,SEB,SAD,SCR,STPL,ACM,CORP,SASPL,SHP,SAH,SCH,SIO,SCA,SAB,SCY,SIP,SIA,SKC,SHN
April'25,3485,10249,17726,40557,15257,0,1264,3601,15245,3706,9800,31944,0,20850,0,39342,10218,12227,42814,86464
May'25,3687,9419,17707,36932,14781,0,1190,2933,12851,3638,8940,32448,0,19683,0,37537,11909,11436,40013,82349
June'25,3391,8710,15864,31688,13223,0,1018,3049,11281,2992,8220,28288,0,16839,0,35972,13190,10015,35427,61745
July'25,2869,5484,14493,30147,12845,0,460,2646,8561,2909,4560,28240,25768,14725,0,32358,12390,8971,30837,62178
Aug'25,2752,6140,13680,26324,11196,0,393,2053,6889,2635,4440,26032,29749,15949,0,31728,12157,7923,31711,65369
September'25,2891,6060,13844,36867,14327,15410,611,2466,9763,3142,7860,29400,14113,24714,0,33104,11837,8969,33565,63908
October'25,2250,7097,9495,26940,11282,14092,797,2779,10306,2723,7950,32049,16986,29842,11447,32291,9577,7854,35547,72851
November'25,2340,5260,8922,19260,8443,13167,761,2297,10119,2048,5220,26651,14814,28687,12762,25974,8800,5768,30739,52266
December'25,2280,2275,9758,23075,7946,9703,339,2202,7651,1998,5076,19608,13483,20253,9214,27711,8983,5155,35319,57275
January'26,2412,2669,9791,26570,8569,8815,160,2349,6286,1998,4036,12138,17977,21261,9772,22537,7800,3233,36920,75108
February'26,2412,5870,12243,28750,8867,12288,527,2961,10810,2577,0,22518,26434,26236,12536,25500,8688,3274,40764,78627
March'26,2980,8118,14945,37631,13787,21774,1345,3750,13198,3110,0,24377,31963,15633,16825,30604,10728,3834,46710,89358"""

FY27_CSV_DATA = """Month,SAG,SCD,SEB,SAD,SCR,STPL,ACM,CORP,SASPL,SHP,SAH,SCH,SIO,SCA,SAB,SCY,SIP,SIA,SKC,SHN
April'26,3245,9079,16958,40627,16232,27557,1742,3696,14271,3704,0,31789,38692,17552,18763,35156,10373,5437,44884,85838
May'26,3120,10500,19259,42113,17607,29168,2089,3900,14504,3946,0,34594,44353,18785,19767,35751,9605,6270,45175,81833"""

# --- INITIAL STATE MANAGEMENT ---
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if "chat_history" not in st.session_state:
    st.session_state["chat_history"] = [
        {"role": "assistant", "content": "Telemetry interface fully online. Ask me about your worst-performing locations, total carbon emissions, or request a complete summary."}
    ]

# 🎨 PREMIUM CSS OVERLAYS
if not st.session_state["authenticated"]:
    st.markdown("""
        <style>
        .stApp {
            background: #030712 !important;
            overflow: hidden;
        }
        div[data-testid="stVerticalBlock"] > div:has(.auth-card-wrap) {
            background: rgba(8, 14, 32, 0.65) !important;
            backdrop-filter: blur(25px);
            -webkit-backdrop-filter: blur(25px);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 24px;
            padding: 40px !important;
            box-shadow: 0 30px 60px rgba(0, 0, 0, 0.7), inset 0 1px 2px rgba(255,255,255,0.1);
            z-index: 10;
            position: relative;
            margin-top: 10px;
        }
        .portal-banner h2 {
            color: #00ffcc !important;
            font-weight: 800 !important;
            text-shadow: 0 0 15px rgba(0, 255, 204, 0.3);
            letter-spacing: 0.5px;
        }
        .portal-banner p {
            color: #94a3b8 !important;
        }
        label {
            color: #cbd5e1 !important;
        }
        </style>
        """, unsafe_allow_html=True)
else:
    st.markdown("""
        <style>
        @keyframes smoothScaleUp {
            from { opacity: 0; transform: translateY(15px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .bubble-wrapper, .stExpander {
            animation: smoothScaleUp 0.5s cubic-bezier(0.16, 1, 0.3, 1) both;
        }

        /* BUBBLE KPI CARD DESIGN */
        .bubble-wrapper {
            display: flex;
            justify-content: space-around;
            align-items: center;
            flex-wrap: wrap;
            gap: 20px;
            margin: 30px 0;
        }
        .kpi-circle-card {
            width: 195px;
            height: 195px;
            background: radial-gradient(circle at 30% 30%, #ffffff, #f8fafc);
            border: 2px solid #e2e8f0;
            border-radius: 50%;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.05);
            transition: transform 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275), box-shadow 0.4s, border-color 0.4s;
            text-align: center;
            padding: 15px;
        }
        .kpi-circle-card:hover {
            transform: translateY(-8px) scale(1.05);
            box-shadow: 0 20px 35px rgba(16, 185, 129, 0.18);
            border-color: #10b981;
        }
        .bubble-title {
            font-size: 11px;
            text-transform: uppercase;
            letter-spacing: 0.6px;
            color: #64748b;
            font-weight: 700;
            margin-bottom: 6px;
        }
        .bubble-value {
            font-size: 17px;
            font-weight: 800;
            color: #0f172a;
            line-height: 1.2;
        }
        </style>
        """, unsafe_allow_html=True)


# 2. Portal Security Wall
if not st.session_state["authenticated"]:
    st.markdown(
        """
        <div style="position:fixed; top:0; left:0; width:100vw; height:100vh; background:#030712; overflow:hidden; z-index:-1; display:flex; justify-content:center; align-items:center;">
            <div id="globeContainer" style="width:500px; height:500px;"></div>
        </div>
        """, unsafe_allow_html=True
    )

    _, col_center, _ = st.columns([1, 1.3, 1])
    with col_center:
        st.markdown('<div class="auth-card-wrap"></div>', unsafe_allow_html=True)
        st.markdown('<div class="portal-banner" style="text-align: center; margin-bottom:20px;"><h2>🌱 Sandhar Energy Portal</h2><p>Ecosystem Identity Verification Matrix</p></div>', unsafe_allow_html=True)
        username = st.text_input("Matrix Operator Key", placeholder="Username ID")
        password = st.text_input("Access Authorization Token", type="password", placeholder="••••••••")
        if st.button("Initialize Energy Workspace", type="primary", use_container_width=True):
            if username == "sandhar" and password == "telemetry2026":
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("System access codes rejected.")
    st.stop()


# 3. Load Datasets via Cached Handlers
@st.cache_data
def load_energy_data_matrices():
    df_m = pd.read_csv(io.StringIO(MASTER_CSV_DATA.strip()))
    df_t = pd.read_csv(io.StringIO(MONTHLY_CSV_DATA.strip()))
    return df_m, df_t

df_master, df_monthly = load_energy_data_matrices()
df_fy27 = pd.read_csv(io.StringIO(FY27_CSV_DATA.strip()))


# --- SIDEBAR CONTROL PANEL ---
st.sidebar.markdown("🔒 **Telemetry Link Stable**")
if st.sidebar.button("Log Out Context"):
    st.session_state["authenticated"] = False
    st.rerun()

st.sidebar.header("🗺️ Application Pages")
app_page = st.sidebar.radio("Navigate Workspace", ["Main Tracking Panel", "FY26-27 Analytics & Horizon Panel"])


# ================= PAGE 2: NEW ANALYTICS PANEL =================
if app_page == "FY26-27 Analytics & Horizon Panel":
    st.title("🚀 FY26-27 Next Horizon Engine")
    st.caption("Active forecasting layers and validation horizons parsed from incoming live execution spreadsheets.")
    st.markdown("---")
    
    st.subheader("📋 Infrastructure Node Matrix Evaluation Ledger (FY26-27 Data Metrics)")
    for idx, row in df_master.iterrows():
        unit_code = str(row['unit']).strip()
        
        apr_series = df_fy27.loc[df_fy27['Month'] == "April'26", unit_code].values
        may_series = df_fy27.loc[df_fy27['Month'] == "May'26", unit_code].values
        
        apr_val = apr_series if len(apr_series) > 0 else 0
        may_val = may_series if len(may_series) > 0 else 0
        
        with st.expander(f"🏢 Node Layer [{unit_code}] — Horizon Status Analysis"):
            col_a, col_b, col_c = st.columns(3)
            col_a.metric("April'26 Yield Log", f"{int(apr_val):,} kWh")
            col_b.metric("May'26 Yield Log", f"{int(may_val):,} kWh")
            
            y_ratio27 = float(row['gen_kwp_27'])
            if y_ratio27 > 3.0:
                col_c.markdown(f"**Generation per KWP (FY26-27)**<br><span style='color:#10b981; font-size:24px; font-weight:bold;'>🟢 {y_ratio27} Yield</span>", unsafe_allow_html=True)
            else:
                col_c.markdown(f"**Generation per KWP (FY26-27)**<br><span style='color:#ef4444; font-size:24px; font-weight:bold;'>🔴 {y_ratio27} Yield</span>", unsafe_allow_html=True)
    st.stop()


# ================= PAGE 1: MAIN TRACKING PANEL =================
st.sidebar.header("🕹️ Selection Filters")
selected_vertical = st.sidebar.selectbox("Business Segment", ["All Segments"] + list(df_master['vertical'].unique()), key="main_vert")
df_filtered = df_master.copy() if selected_vertical == "All Segments" else df_master[df_master['vertical'] == selected_vertical].copy()

target_month = st.sidebar.select_slider("Select Target Tracking Month (FY25-26)", options=list(df_monthly['Month']))

# 🟢 1. BUBBLE KPI CARDS 
total_grid = df_filtered['grid_mvah'].sum()
total_mit = df_filtered['mitigation'].sum()
total_emi = df_filtered['emission'].sum()

st.markdown("### 📊 Metrics Summary Grid")
col_metric_1, col_metric_2, col_metric_3 = st.columns(3)
col_metric_1.metric("⚡ Total Grid Sourced", f"{total_grid:,.1f} MVAh")
col_metric_2.metric("🌱 Carbon Offset", f"{int(total_mit):,} MT CO₂")
col_metric_3.metric("🏭 Gross Footprint", f"{int(total_emi):,} MT CO₂")


# ================= AUTOMATED REPORT GENERATION ENGINE =================
def generate_excel_report(dataframe):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ecosystem Data"
    
    ws.sheet_view.showGridLines = True
    
    title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    ws.merge_cells("A1:G1")
    ws["A1"] = "SANDHAR ECOSYSTEM DATA MANAGEMENT REPORT"
    ws["A1"].font = font_title
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions.height = 35
    
    cols = ["Vertical Segment", "Plant Node", "Location", "Grid (MVAh)", "Mitigation (MT)", "Emissions (MT)", "Yield Ratio"]
    for idx, text in enumerate(cols, 1):
        cell = ws.cell(row=3, column=idx, value=text)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions.height = 24
    
    for r_idx, row in dataframe.reset_index().iterrows():
        r = r_idx + 4
        ws.cell(row=r, column=1, value=row['vertical']).font = font_data
        ws.cell(row=r, column=2, value=row['unit']).font = font_data
        ws.cell(row=r, column=3, value=row['location']).font = font_data
        
        c4 = ws.cell(row=r, column=4, value=row['grid_mvah'])
        c4.number_format = "#,##0.00"
        c4.font = font_data
        
        c5 = ws.cell(row=r, column=5, value=row['mitigation'])
        c5.number_format = "#,##0"
        c5.font = font_data
        
        c6 = ws.cell(row=r, column=6, value=row['emission'])
        c6.number_format = "#,##0"
        c6.font = font_data
        
        c7 = ws.cell(row=r, column=7, value=row['generation_per_kwp'])
        c7.number_format = "0.00"
        c7.font = font_data
        
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col.column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output)
    return output.getvalue()

excel_data = generate_excel_report(df_filtered)
st.download_button(
    label="📥 Download Formatted Ecosystem Executive Report (.xlsx)",
    data=excel_data,
    file_name=f"Sandhar_Energy_Report_{target_month}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

st.markdown("---")

# 📊 2. DYNAMIC VISUALIZATION GRAPH BLOCK
st.subheader("📊 Dynamic Environmental Performance & Fleet Generation Metrics")
col_g1, col_g2 = st.columns(2)

with col_g1:
    fig_bar = px.bar(
        df_filtered, 
        x='unit', 
        y=['mitigation', 'emission'],
        barmode='group',
        title="Carbon Offset (Mitigation) vs Gross Footprint by Operational Node",
        labels={'value': 'Metric Tons (CO₂)', 'unit': 'Plant Node Code'},
        color_discrete_sequence=['#10b981', '#ef4444']
    )
    fig_bar.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', legend_title_text='Metrics')
    st.plotly_chart(fig_bar, use_container_width=True)

with col_g2:
    fig_scatter = px.scatter(
        df_filtered,
        x='grid_mvah',
        y='generation_per_kwp',
        size='unit_lost_inefficiency',
        color='vertical',
        hover_name='unit',
        title='Generation Ratio vs Grid Sourcing (Bubble size = Inefficiency Losses)',
        labels={'grid_mvah': 'Grid Sourced (MVAh)', 'generation_per_kwp': 'Gen/KWP Ratio'}
    )
    fig_scatter.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
    st.plotly_chart(fig_scatter, use_container_width=True)

st.markdown("---")

# 🗺️ 3. INTERACTIVE FOLLIUM MAP EMBED
st.subheader("🗺️ Enterprise Infrastructure Geolocation Node Overlay")
if not df_filtered.empty:
    avg_lat = df_filtered['lat'].mean()
    avg_lon = df_filtered['lon'].mean()
    
    # 🟢 FIXED: Re-added fully intact configuration initialization tuple wrapper safely
    m = folium.Map(location=[avg_lat, avg_lon], zoom_start=5, tiles="CartoDB positron")
    
    for _, marker_row in df_filtered.iterrows():
        popup_html = f"""
        <div style='font-family: Arial, sans-serif; font-size:12px; line-height: 1.4;'>
            <strong>Node Code:</strong> {marker_row['unit']}<br>
            <strong>Location:</strong> {marker
